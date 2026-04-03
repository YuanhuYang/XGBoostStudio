"""
生成测试数据集脚本
运行：cd server && uv run python tests/create_fixtures.py
"""
import random
import math
import csv
from pathlib import Path

FIXTURES_DIR = Path(__file__).parent / "fixtures"
FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

random.seed(42)


def rand_normal(mu=0.0, sigma=1.0):
    """Box-Muller 正态分布（避免依赖 numpy）"""
    u1 = random.random() or 1e-10
    u2 = random.random()
    return mu + sigma * math.sqrt(-2 * math.log(u1)) * math.cos(2 * math.pi * u2)


# ── Titanic 训练集（891行×12列，二分类）────────────────────────────────────────
def create_titanic_train():
    """生成 Titanic 训练集（891行×12列，二分类）并写入 CSV 文件。"""
    headers = ["PassengerId", "Survived", "Pclass", "Name", "Sex", "Age",
               "SibSp", "Parch", "Ticket", "Fare", "Cabin", "Embarked"]
    rows = []
    for i in range(1, 892):
        pclass = random.choice([1, 2, 3])
        sex = random.choice(["male", "female"])
        age = round(max(1, rand_normal(30, 14)), 1) if random.random() > 0.2 else ""
        sibsp = random.choice([0, 0, 0, 1, 1, 2, 3])
        parch = random.choice([0, 0, 0, 1, 2])
        fare = round(max(5, rand_normal(32, 49)), 2)
        cabin = f"C{random.randint(10,150)}" if random.random() > 0.7 else ""
        embarked = random.choice(["S", "S", "S", "C", "Q"]) if random.random() > 0.02 else ""
        survived = 1 if (sex == "female" and random.random() > 0.25) or \
                       (sex == "male" and pclass == 1 and random.random() > 0.6) or \
                       (sex == "male" and random.random() > 0.8) else 0
        rows.append([i, survived, pclass, f"Passenger_{i}", sex, age,
                     sibsp, parch, f"T{i:05d}", fare, cabin, embarked])
    with open(FIXTURES_DIR / "titanic_train.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"✓ titanic_train.csv ({len(rows)} rows)")


# ── Titanic 测试集（418行×11列，无目标列）─────────────────────────────────────
def create_titanic_test():
    """生成 Titanic 测试集（418行×11列，无目标列）并写入 CSV 文件。"""
    headers = ["PassengerId", "Pclass", "Name", "Sex", "Age",
               "SibSp", "Parch", "Ticket", "Fare", "Cabin", "Embarked"]
    rows = []
    for i in range(892, 1310):
        pclass = random.choice([1, 2, 3])
        sex = random.choice(["male", "female"])
        age = round(max(1, rand_normal(30, 14)), 1) if random.random() > 0.2 else ""
        sibsp = random.choice([0, 0, 1, 1, 2])
        parch = random.choice([0, 0, 1])
        fare = round(max(5, rand_normal(32, 49)), 2)
        cabin = f"C{random.randint(10,150)}" if random.random() > 0.7 else ""
        embarked = random.choice(["S", "C", "Q"])
        rows.append([i, pclass, f"Passenger_{i}", sex, age,
                     sibsp, parch, f"T{i:05d}", fare, cabin, embarked])
    with open(FIXTURES_DIR / "titanic_test.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"✓ titanic_test.csv ({len(rows)} rows)")


# ── Boston Housing（506行×14列，回归）─────────────────────────────────────────
def create_boston_housing():
    """生成 Boston Housing 数据集（506行×14列，回归）并写入 CSV 文件。"""
    headers = ["CRIM", "ZN", "INDUS", "CHAS", "NOX", "RM", "AGE",
               "DIS", "RAD", "TAX", "PTRATIO", "B", "LSTAT", "MEDV"]
    rows = []
    for _ in range(506):
        crim = round(max(0.01, rand_normal(3.6, 8.6)), 4)
        zn = round(max(0, rand_normal(11, 23)), 1)
        indus = round(max(0.5, rand_normal(11, 7)), 2)
        chas = random.choice([0, 0, 0, 0, 1])
        nox = round(max(0.3, min(0.9, rand_normal(0.55, 0.12))), 3)
        rm = round(max(3, min(9, rand_normal(6.28, 0.7))), 3)
        age = round(max(2, min(100, rand_normal(68, 28))), 1)
        dis = round(max(1, rand_normal(3.8, 2.1)), 4)
        rad = random.choice([1, 2, 3, 4, 5, 6, 7, 8, 24])
        tax = random.choice([193, 222, 264, 307, 330, 384, 403, 432, 666])
        ptratio = round(max(12, min(22, rand_normal(18.5, 2.2))), 1)
        b = round(max(0, min(400, rand_normal(356, 91))), 2)
        lstat = round(max(1, rand_normal(12.6, 7)), 2)
        medv = round(max(5, min(50, rand_normal(22.5, 9.2))), 1)
        rows.append([crim, zn, indus, chas, nox, rm, age, dis, rad, tax, ptratio, b, lstat, medv])
    with open(FIXTURES_DIR / "boston_housing.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"✓ boston_housing.csv ({len(rows)} rows)")


# ── Iris（150行×5列，多分类）──────────────────────────────────────────────────
def create_iris():
    """生成 Iris 数据集（150行×5列，多分类）并写入 CSV 文件。"""
    headers = ["sepal_length", "sepal_width", "petal_length", "petal_width", "species"]
    specs = [
        ("setosa", 5.0, 0.35, 3.4, 0.38, 1.46, 0.17, 0.24, 0.11),
        ("versicolor", 5.94, 0.52, 2.77, 0.31, 4.26, 0.47, 1.33, 0.20),
        ("virginica", 6.59, 0.64, 2.97, 0.32, 5.55, 0.55, 2.03, 0.27),
    ]
    rows = []
    for name, sl_mu, sl_s, sw_mu, sw_s, pl_mu, pl_s, pw_mu, pw_s in specs:
        for _ in range(50):
            rows.append([
                round(max(4, rand_normal(sl_mu, sl_s)), 1),
                round(max(2, rand_normal(sw_mu, sw_s)), 1),
                round(max(1, rand_normal(pl_mu, pl_s)), 1),
                round(max(0.1, rand_normal(pw_mu, pw_s)), 1),
                name,
            ])
    random.shuffle(rows)
    with open(FIXTURES_DIR / "iris.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"✓ iris.csv ({len(rows)} rows)")


# ── Large 100k（10万行×20列，性能测试）───────────────────────────────────────
def create_large_100k():
    """生成大规模数据集（10万行×20列，性能测试）并写入 CSV 文件。"""
    headers = [f"feature_{i}" for i in range(1, 21)] + ["target"]
    with open(FIXTURES_DIR / "large_100k.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for _ in range(100000):
            row = [round(rand_normal(0, 1), 4) for _ in range(20)]
            target = 1 if sum(row[:5]) > 0 else 0
            row.append(target)
            writer.writerow(row)
    print("✓ large_100k.csv (100000 rows)")


# ── Missing Heavy（500行，缺失率约30%）───────────────────────────────────────
def create_missing_heavy():
    """生成含大量缺失值的数据集（500行，缺失率约30%）并写入 CSV 文件。"""
    headers = [f"col_{i}" for i in range(1, 11)] + ["target"]
    rows = []
    for _ in range(500):
        row = []
        for _ in range(10):
            if random.random() < 0.30:
                row.append("")
            else:
                row.append(round(rand_normal(0, 1), 3))
        row.append(random.choice([0, 1]))
        rows.append(row)
    with open(FIXTURES_DIR / "missing_heavy.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"✓ missing_heavy.csv ({len(rows)} rows)")


# ── Duplicate Rows（200行含50条重复行）───────────────────────────────────────
def create_duplicate_rows():
    """生成含重复行的数据集（200行含50条重复行）并写入 CSV 文件。"""
    headers = ["id", "name", "value", "category", "target"]
    base = []
    for i in range(150):
        base.append([i, f"item_{i}", round(rand_normal(50, 15), 2),
                     random.choice(["A", "B", "C"]), random.choice([0, 1])])
    duplicates = random.choices(base, k=50)
    all_rows = base + duplicates
    random.shuffle(all_rows)
    with open(FIXTURES_DIR / "duplicate_rows.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(all_rows)
    print(f"✓ duplicate_rows.csv ({len(all_rows)} rows, 50 duplicates)")


# ── Multisheet Excel（3个Sheet）──────────────────────────────────────────────
def create_multisheet_excel():
    """生成多 Sheet 的 Excel 文件（Sales/Customers/Products 三个工作表）。"""
    try:
        import openpyxl
        wb = openpyxl.Workbook()

        # Sheet1: Sales
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["date", "product", "region", "sales", "profit"])
        for i in range(100):
            ws1.append([f"2024-{random.randint(1,12):02d}-{random.randint(1,28):02d}",
                        random.choice(["A", "B", "C"]),
                        random.choice(["North", "South", "East", "West"]),
                        round(rand_normal(5000, 2000), 2),
                        round(rand_normal(800, 300), 2)])

        # Sheet2: Customers
        ws2 = wb.create_sheet("Customers")
        ws2.append(["customer_id", "age", "income", "score", "churned"])
        for i in range(200):
            ws2.append([i + 1, random.randint(18, 75),
                        round(rand_normal(50000, 20000), 0),
                        round(rand_normal(700, 100), 0),
                        random.choice([0, 1])])

        # Sheet3: Products
        ws3 = wb.create_sheet("Products")
        ws3.append(["product_id", "category", "price", "rating", "in_stock"])
        for i in range(80):
            ws3.append([f"P{i+1:03d}",
                        random.choice(["Electronics", "Clothing", "Food", "Books"]),
                        round(rand_normal(100, 80), 2),
                        round(rand_normal(3.8, 0.8), 1),
                        random.choice([True, False])])

        wb.save(FIXTURES_DIR / "multisheet.xlsx")
        print("✓ multisheet.xlsx (3 sheets)")
    except ImportError:
        print("⚠ openpyxl 未安装，跳过 multisheet.xlsx 生成")


if __name__ == "__main__":
    print("正在生成测试数据集...")
    create_titanic_train()
    create_titanic_test()
    create_boston_housing()
    create_iris()
    create_large_100k()
    create_missing_heavy()
    create_duplicate_rows()
    create_multisheet_excel()
    print("\n所有测试数据集生成完毕！")
    print(f"位置：{FIXTURES_DIR}")
